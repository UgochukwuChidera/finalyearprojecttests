import csv
import json
from pathlib import Path


class DataExporter:
    """
    Exports a structured output document (from OutputStructurer) to disk
    in JSON, CSV, and/or XLSX formats.

    JSON is always the primary canonical format. CSV and XLSX are auxiliary
    exports that flatten the `data` dict into two-column tables.
    """

    # ── JSON ───────────────────────────────────────────────────────────────────

    def export_json(self, structured: dict, output_path: str) -> str:
        """Write the full structured document as pretty-printed JSON."""
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8") as fh:
            json.dump(structured, fh, indent=2, default=str, ensure_ascii=False)
        return str(path)

    # ── CSV ────────────────────────────────────────────────────────────────────

    def export_csv(self, structured: dict, output_path: str) -> str:
        """
        Write the `data` dict as a two-column CSV (field, value).
        Also appends form_id and template_id as metadata rows.
        """
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(["field", "value"])
            writer.writerow(["_form_id",     structured.get("form_id", "")])
            writer.writerow(["_template_id", structured.get("template_id", "")])
            writer.writerow(["_processed_at", structured.get("processed_at", "")])
            for k, v in structured.get("data", {}).items():
                writer.writerow([k, v])
        return str(path)

    # ── XLSX ───────────────────────────────────────────────────────────────────

    def export_xlsx(self, structured: dict, output_path: str) -> str:
        """
        Write the `data` dict as a formatted XLSX workbook.
        Requires the `openpyxl` package.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError(
                "openpyxl is required for XLSX export. "
                "Install it with: pip install openpyxl"
            )

        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extracted Data"

        # Header row
        header_fill = PatternFill("solid", fgColor="1A1A2E")
        header_font = Font(bold=True, color="FFFFFF")
        for col, label in enumerate(["Field", "Value"], start=1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.fill = header_fill
            cell.font = header_font

        # Metadata rows
        meta = [
            ("_form_id",      structured.get("form_id", "")),
            ("_template_id",  structured.get("template_id", "")),
            ("_processed_at", structured.get("processed_at", "")),
        ]
        row = 2
        for k, v in meta:
            ws.cell(row=row, column=1, value=k)
            ws.cell(row=row, column=2, value=str(v))
            row += 1

        # Data rows
        for k, v in structured.get("data", {}).items():
            ws.cell(row=row, column=1, value=k)
            ws.cell(row=row, column=2, value=str(v))
            row += 1

        # Auto-fit column widths (approximate)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 40

        wb.save(str(path))
        return str(path)

    # ── Convenience: export all formats ───────────────────────────────────────

    def export_all(self, structured: dict, base_path: str) -> dict[str, str]:
        """
        Export JSON, CSV, and XLSX to *base_path*.json / .csv / .xlsx.

        Returns a mapping of format → output file path.
        """
        base = str(Path(base_path).with_suffix(""))
        return {
            "json": self.export_json(structured, base + ".json"),
            "csv":  self.export_csv( structured, base + ".csv"),
            "xlsx": self.export_xlsx(structured, base + ".xlsx"),
        }
